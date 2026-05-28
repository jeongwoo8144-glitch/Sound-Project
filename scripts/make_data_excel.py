import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from collections import defaultdict

base = '/mnt/c/Users/Daniel Park/Desktop/SoundPJ-rebuilt/data/processed'
out_path = '/mnt/c/Users/Daniel Park/Desktop/SoundPJ-rebuilt/results/data_file_counts.xlsx'

# Collect all file counts (skip embeddings folder)
data = {}
for snr in sorted(os.listdir(base)):
    if snr == 'embeddings':
        continue
    snr_path = os.path.join(base, snr)
    if not os.path.isdir(snr_path):
        continue
    for fold in sorted(os.listdir(snr_path)):
        fold_path = os.path.join(snr_path, fold)
        if not os.path.isdir(fold_path):
            continue
        for cls in sorted(os.listdir(fold_path)):
            cls_path = os.path.join(fold_path, cls)
            if not os.path.isdir(cls_path):
                continue
            count = len([f for f in os.listdir(cls_path) if os.path.isfile(os.path.join(cls_path, f))])
            data[(snr, fold, cls)] = count

# Sort keys
snr_order_raw = sorted(data.keys(), key=lambda x: x[0])
snrs = sorted(set(k[0] for k in data.keys()))
folds = sorted(set(k[1] for k in data.keys()), key=lambda x: int(x.replace('fold','')))
classes = sorted(set(k[2] for k in data.keys()))

print("SNRs:", snrs)
print("Folds:", folds)
print("Classes:", classes)
print("Total entries:", len(data))

# Style helpers
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
SUB_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
NORM_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

SNR_COLORS = {
    "clean":       "E8F5E9",
    "snr_+10dB":   "FFF9C4",
    "snr_+5dB":    "FFE0B2",
    "snr_+0dB":    "FCE4EC",
    "snr_-5dB":    "EDE7F6",
}
CLASS_COLORS = {
    "background": "E3F2FD",
    "car_horn":   "FFF3E0",
    "siren":      "FCE4EC",
}

FOLD_ROLE = {}
for f in folds:
    n = int(f.replace("fold", ""))
    if n <= 8:
        FOLD_ROLE[f] = "Train"
    elif n == 9:
        FOLD_ROLE[f] = "Val"
    else:
        FOLD_ROLE[f] = "Test"

wb = openpyxl.Workbook()

# ── Sheet 1: 상세 데이터 ──────────────────────────────────────────
ws1 = wb.active
ws1.title = "상세 데이터"

headers = ["SNR 조건", "Fold", "역할", "클래스", "파일 수"]
col_widths = [16, 8, 8, 14, 10]

for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=1, column=ci, value=h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = CENTER
    cell.border = BORDER
    ws1.column_dimensions[get_column_letter(ci)].width = w

row = 2
for snr in snrs:
    for fold in folds:
        for cls in classes:
            val = data.get((snr, fold, cls), 0)
            if val == 0:
                continue
            row_data = [snr, fold, FOLD_ROLE[fold], cls, val]
            for ci, v in enumerate(row_data, 1):
                cell = ws1.cell(row=row, column=ci, value=v)
                cell.font = NORM_FONT
                cell.border = BORDER
                if ci == 5:
                    cell.alignment = CENTER
                else:
                    cell.alignment = LEFT
                if ci == 1:
                    cell.fill = PatternFill("solid", fgColor=SNR_COLORS.get(snr, "FFFFFF"))
                elif ci == 4:
                    cell.fill = PatternFill("solid", fgColor=CLASS_COLORS.get(cls, "FFFFFF"))
                elif row % 2 == 0:
                    cell.fill = ALT_FILL
            row += 1

ws1.freeze_panes = "A2"
ws1.auto_filter.ref = "A1:E1"

# ── Sheet 2: SNR별 요약 ──────────────────────────────────────────
ws2 = wb.create_sheet("SNR별 요약")

# Header row: SNR | background | car_horn | siren | 합계
h2 = ["SNR 조건"] + classes + ["합계"]
for ci, h in enumerate(h2, 1):
    cell = ws2.cell(row=1, column=ci, value=h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = CENTER
    cell.border = BORDER

ws2.column_dimensions["A"].width = 16
for ci in range(2, len(h2) + 1):
    ws2.column_dimensions[get_column_letter(ci)].width = 14

for ri, snr in enumerate(snrs, 2):
    row_vals = [snr]
    total = 0
    for cls in classes:
        s = sum(data.get((snr, f, cls), 0) for f in folds)
        row_vals.append(s)
        total += s
    row_vals.append(total)
    for ci, v in enumerate(row_vals, 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        cell.font = NORM_FONT
        cell.border = BORDER
        cell.alignment = CENTER if ci > 1 else LEFT
        if ci == 1:
            cell.fill = PatternFill("solid", fgColor=SNR_COLORS.get(snr, "FFFFFF"))
        elif ri % 2 == 0:
            cell.fill = ALT_FILL

# Total row
tr = len(snrs) + 2
ws2.cell(row=tr, column=1, value="전체 합계").font = Font(name="Arial", bold=True, size=10)
ws2.cell(row=tr, column=1).border = BORDER
ws2.cell(row=tr, column=1).alignment = CENTER
for ci, cls in enumerate(classes, 2):
    s = sum(data.get((snr, f, cls), 0) for snr in snrs for f in folds)
    cell = ws2.cell(row=tr, column=ci, value=s)
    cell.font = Font(name="Arial", bold=True, size=10)
    cell.fill = PatternFill("solid", fgColor="BDD7EE")
    cell.border = BORDER
    cell.alignment = CENTER
grand = sum(data.values())
cell = ws2.cell(row=tr, column=len(h2), value=grand)
cell.font = Font(name="Arial", bold=True, size=10)
cell.fill = PatternFill("solid", fgColor="1F4E79")
cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
cell.border = BORDER
cell.alignment = CENTER

# ── Sheet 3: Fold별 요약 ──────────────────────────────────────────
ws3 = wb.create_sheet("Fold별 요약")

ROLE_FILL = {"Train": "E8F5E9", "Val": "FFF9C4", "Test": "FCE4EC"}

h3 = ["Fold", "역할"] + classes + ["합계"]
for ci, h in enumerate(h3, 1):
    cell = ws3.cell(row=1, column=ci, value=h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = CENTER
    cell.border = BORDER

ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 8
for ci in range(3, len(h3) + 1):
    ws3.column_dimensions[get_column_letter(ci)].width = 14

for ri, fold in enumerate(folds, 2):
    role = FOLD_ROLE[fold]
    row_vals = [fold, role]
    total = 0
    for cls in classes:
        s = sum(data.get((snr, fold, cls), 0) for snr in snrs)
        row_vals.append(s)
        total += s
    row_vals.append(total)
    for ci, v in enumerate(row_vals, 1):
        cell = ws3.cell(row=ri, column=ci, value=v)
        cell.font = NORM_FONT
        cell.border = BORDER
        cell.alignment = CENTER
        cell.fill = PatternFill("solid", fgColor=ROLE_FILL.get(role, "FFFFFF"))

# Legend
ws3.cell(row=len(folds) + 3, column=1, value="범례:").font = Font(name="Arial", bold=True, size=10)
for i, (role, color) in enumerate(ROLE_FILL.items()):
    c = ws3.cell(row=len(folds) + 3, column=i + 2, value=role)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = NORM_FONT
    c.alignment = CENTER
    c.border = BORDER

os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print("Saved:", out_path)
print("Grand total files:", sum(data.values()))
